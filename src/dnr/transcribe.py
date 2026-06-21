"""Transcription providers (M4) — dnr is the storage/trust layer, not the model.

The transcript is an *input*, produced by whoever is best placed:
  - the **calling agent** (already a vision/multimodal LLM) -> supplies the verbatim
    text via ``dnr record`` (primary path for visual docs / images / scans),
  - a **local model or extractor** (Whisper for audio, ``text-extract`` for born-digital
    PDFs / Office files),
  - optionally a hosted API.

Providers here are the *local* ones. The agent path lives in ``ingest.record_supplied``.
"""
from __future__ import annotations

import unicodedata
from dataclasses import dataclass, field


@dataclass
class TranscriptResult:
    text: str
    method: str
    transcriber: str
    lang: str | None = None
    confidence: float | None = None
    segments: list | None = field(default=None)


def detect_lang(text: str | None) -> str | None:
    """Cheap, deterministic, dependency-free script heuristic (ko / zh / ja / en)."""
    if not text:
        return None
    han = cjk = kana = latin = 0
    for c in text:
        o = ord(c)
        if 0xAC00 <= o <= 0xD7A3 or 0x1100 <= o <= 0x11FF:
            han += 1
        elif 0x3040 <= o <= 0x30FF:
            kana += 1
        elif 0x4E00 <= o <= 0x9FFF:
            cjk += 1
        elif c.isascii() and c.isalpha():
            latin += 1
    if han and han >= cjk:
        return "ko"
    if kana:
        return "ja"
    if cjk:
        return "zh"
    if latin:
        return "en"
    return None


def readability_score(text: str | None) -> float:
    """Cheap script-aware readability score in [0, 1]."""
    if not text or len(text.strip()) < 3:
        return 0.0
    s = text.strip()
    ok_ascii = " \t\n\r.,;:!?()[]{}'\"/\\-–—…%₩$+*=&@#·"
    readable = 0
    for c in s:
        o = ord(c)
        if 0xAC00 <= o <= 0xD7A3 or 0x4E00 <= o <= 0x9FFF or 0x3040 <= o <= 0x30FF:  # Hangul / CJK / kana
            readable += 1
        elif c.isascii() and (c.isalnum() or c in ok_ascii):
            readable += 1
    return readable / len(s)


def is_low_quality(text: str | None) -> bool:
    """Cheap heuristic: is this transcript empty or **garbled** (e.g. EUC-KR/CP949 decoded as
    Latin-1 → mojibake)? We don't try to *fix* it — that's the vision/`dnr record` path's job;
    we just flag it so it isn't silently trusted."""
    return readability_score(text) < 0.55  # mostly unreadable bytes -> mojibake / garbage


def text_extract(path) -> TranscriptResult:
    """Born-digital PDF -> embedded text layer, NFC-normalized. Lossless, no model.

    PyMuPDF is tried first because it is often better on CJK/Korean PDFs. pypdf remains a
    fallback so older environments and odd PDFs still have a second path.
    """
    candidates: list[tuple[str, str]] = []
    errors: list[str] = []

    try:
        import fitz  # PyMuPDF

        with fitz.open(str(path)) as doc:
            parts = [page.get_text("text") or "" for page in doc]
        candidates.append(("pymupdf", "\n\f\n".join(unicodedata.normalize("NFC", p) for p in parts).strip()))
    except Exception as exc:
        errors.append(f"pymupdf: {exc}")

    try:
        from pypdf import PdfReader

        parts = []
        for page in PdfReader(str(path)).pages:
            parts.append(unicodedata.normalize("NFC", page.extract_text() or ""))
        candidates.append(("pypdf", "\n\f\n".join(parts).strip()))
    except Exception as exc:
        errors.append(f"pypdf: {exc}")

    if not candidates:
        raise RuntimeError("PDF text extraction failed (" + "; ".join(errors) + ")")

    # Prefer readable text; tie-break by length. This keeps PyMuPDF first for comparable output.
    transcriber, text = max(
        candidates,
        key=lambda item: (readability_score(item[1]), len(item[1] or ""), -candidates.index(item)),
    )
    return TranscriptResult(text=text, method="text-extract", transcriber=transcriber)


DEFAULT_WHISPER_MODEL = "small"


def whisper_transcribe(path, model_size: str = DEFAULT_WHISPER_MODEL) -> TranscriptResult:
    """Local ASR for audio via faster-whisper. Verbatim by construction; timestamps."""
    try:
        from faster_whisper import WhisperModel
    except ImportError as exc:  # optional heavy dep + model download
        raise RuntimeError(
            "the 'whisper' provider needs the audio extras: install with "
            "`pip install 'donotreadagain[audio]'`, `pipx inject donotreadagain faster-whisper`, "
            "or run via `uvx --from 'donotreadagain[audio]' dnr ...`. If decoding fails, install "
            "ffmpeg too. You can also supply the transcript via `dnr record`."
        ) from exc

    model = WhisperModel(model_size, compute_type="int8")
    segments, info = model.transcribe(str(path))
    segs, parts = [], []
    for s in segments:
        t = s.text.strip()
        segs.append({"t": round(s.start, 3), "text": t})
        parts.append(t)
    return TranscriptResult(
        text="\n".join(parts),
        method="asr",
        transcriber=f"faster-whisper-{model_size}",
        lang=getattr(info, "language", None),
        segments=segs,
    )


def docx_extract(path) -> TranscriptResult:
    """Born-digital Word document -> paragraph text via python-docx. Local, no model."""
    import docx

    document = docx.Document(str(path))
    text = "\n".join(p.text for p in document.paragraphs)
    return TranscriptResult(text=unicodedata.normalize("NFC", text), method="text-extract", transcriber="python-docx")


def _xlsx_cell_text(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def xlsx_extract(path) -> TranscriptResult:
    """Spreadsheet workbook -> sheet/row text via openpyxl. Local, no model."""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts: list[str] = []
    for ws in wb.worksheets:
        parts.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = [_xlsx_cell_text(v) for v in row]
            while cells and not cells[-1]:
                cells.pop()
            if any(cells):
                parts.append("\t".join(cells))
        parts.append("")
    text = unicodedata.normalize("NFC", "\n".join(parts).strip())
    return TranscriptResult(text=text, method="text-extract", transcriber="openpyxl")


#: name -> local provider callable. Local OCR / vision register here later.
REGISTRY = {
    "text-extract": text_extract,
    "whisper": whisper_transcribe,
    "docx": docx_extract,
    "xlsx": xlsx_extract,
}


def get(name: str):
    if name not in REGISTRY:
        raise ValueError(
            f"unknown transcriber '{name}'. local: {sorted(REGISTRY)}; "
            f"or supply the transcript yourself via `dnr record`."
        )
    return REGISTRY[name]
